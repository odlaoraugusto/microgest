"""
Service do módulo Relatórios (Sprint 10).

Gera arquivos em memória (BytesIO) para exportação - nunca grava nada
em disco no servidor. Cada método devolve os bytes prontos para serem
enviados como resposta HTTP (StreamingResponse) pelo Router.
"""
import io
import os
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Image,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy.orm import Session

from app.repositories.paciente_repository import PacienteRepository
from app.repositories.solicitacao_repository import SolicitacaoRepository
from app.services.ccih_service import CCIHService
from app.services.cultura_service import CulturaService

CABECALHO_FILL = PatternFill(start_color="0F4C81", end_color="0F4C81", fill_type="solid")
CABECALHO_FONT = Font(color="FFFFFF", bold=True)

LOGO_PATH = os.path.join(os.path.dirname(__file__), "..", "assets", "logo.png")


def _estilizar_cabecalho(ws, colunas: list[str]) -> None:
    ws.append(colunas)
    for cell in ws[1]:
        cell.fill = CABECALHO_FILL
        cell.font = CABECALHO_FONT
    for i, coluna in enumerate(colunas, start=1):
        ws.column_dimensions[chr(64 + i)].width = max(18, len(coluna) + 4)


class RelatorioService:
    def __init__(self, db: Session):
        self.db = db
        self.paciente_repository = PacienteRepository(db)
        self.solicitacao_repository = SolicitacaoRepository(db)
        self.ccih_service = CCIHService(db)
        self.cultura_service = CulturaService(db)

    def gerar_excel_pacientes(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Pacientes"
        _estilizar_cabecalho(
            ws, ["Prontuário", "Nome", "Setor", "Leito", "Status", "Cadastrado em"]
        )

        pacientes, _ = self.paciente_repository.search(termo=None, skip=0, limit=10_000)
        for p in pacientes:
            ws.append(
                [
                    p.prontuario,
                    p.nome,
                    p.setor or "",
                    p.leito or "",
                    p.status_internacao.value,
                    p.created_at.strftime("%d/%m/%Y %H:%M"),
                ]
            )

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def gerar_excel_solicitacoes(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Solicitações"
        _estilizar_cabecalho(
            ws,
            [
                "Paciente",
                "Prontuário",
                "Material",
                "Origem",
                "Prioridade",
                "Status",
                "Data da Solicitação",
            ],
        )

        solicitacoes, _ = self.solicitacao_repository.search(skip=0, limit=10_000)
        for s in solicitacoes:
            ws.append(
                [
                    s.paciente.nome if s.paciente else "",
                    s.paciente.prontuario if s.paciente else "",
                    s.material,
                    s.origem or "",
                    s.prioridade.value,
                    s.status.value,
                    s.data_solicitacao.strftime("%d/%m/%Y"),
                ]
            )

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def gerar_excel_culturas_parciais(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados Parciais"
        _estilizar_cabecalho(
            ws,
            [
                "Paciente",
                "Prontuário",
                "Material",
                "Resultado Atual",
                "Microrganismo(s)",
                "Previsão de Liberação",
                "Pendência",
            ],
        )

        itens = self.cultura_service.resultados_parciais()
        for cultura, pendencia in itens:
            paciente = cultura.solicitacao.paciente if cultura.solicitacao else None
            nomes_micro = (
                ", ".join(cm.microrganismo.nome for cm in cultura.microrganismos)
                if cultura.microrganismos
                else "—"
            )
            previsao = (
                cultura.previsao_liberacao.strftime("%d/%m/%Y")
                if cultura.previsao_liberacao
                else "—"
            )
            ws.append(
                [
                    paciente.nome if paciente else "",
                    paciente.prontuario if paciente else "",
                    cultura.solicitacao.material if cultura.solicitacao else "",
                    cultura.resultado.value,
                    nomes_micro,
                    previsao,
                    pendencia,
                ]
            )

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def gerar_pdf_ccih(self, data_inicio: date | None, data_fim: date | None) -> bytes:
        indicadores = self.ccih_service.indicadores(data_inicio, data_fim)
        styles = getSampleStyleSheet()

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=2 * cm,
            rightMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )

        cabecalho_texto = [
            Paragraph("MicroGest — Relatório CCIH", styles["Title"]),
            Paragraph(
                f"Período: {indicadores.periodo_inicio.strftime('%d/%m/%Y')} a "
                f"{indicadores.periodo_fim.strftime('%d/%m/%Y')}",
                styles["Normal"],
            ),
            Paragraph(
                f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                styles["Normal"],
            ),
        ]

        if os.path.exists(LOGO_PATH):
            cabecalho = Table(
                [[Image(LOGO_PATH, width=1.6 * cm, height=1.6 * cm), cabecalho_texto]],
                colWidths=[2.2 * cm, None],
            )
            cabecalho.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("LEFTPADDING", (0, 0), (0, 0), 0),
                        ("LEFTPADDING", (1, 0), (1, 0), 10),
                    ]
                )
            )
            elementos = [cabecalho, Spacer(1, 0.8 * cm)]
        else:
            elementos = [*cabecalho_texto, Spacer(1, 0.8 * cm)]

        resumo_dados = [
            ["Solicitações no período", str(indicadores.total_solicitacoes)],
            ["Culturas positivas", str(indicadores.total_culturas_positivas)],
            ["Taxa de positividade", f"{indicadores.taxa_positividade}%"],
        ]
        elementos.append(Paragraph("Resumo Geral", styles["Heading2"]))
        elementos.append(_tabela(resumo_dados, ["Indicador", "Valor"]))
        elementos.append(Spacer(1, 0.6 * cm))

        elementos.append(Paragraph("Distribuição por Setor", styles["Heading2"]))
        if indicadores.distribuicao_por_setor:
            dados_setor = [[d.setor, str(d.total_positivas)] for d in indicadores.distribuicao_por_setor]
            elementos.append(_tabela(dados_setor, ["Setor", "Culturas Positivas"]))
        else:
            elementos.append(Paragraph("Sem dados no período.", styles["Normal"]))
        elementos.append(Spacer(1, 0.6 * cm))

        elementos.append(Paragraph("Perfil Microbiológico", styles["Heading2"]))
        if indicadores.perfil_microbiologico:
            dados_perfil = [
                [p.microrganismo, str(p.quantidade), f"{p.percentual}%"]
                for p in indicadores.perfil_microbiologico
            ]
            elementos.append(_tabela(dados_perfil, ["Microrganismo", "Quantidade", "% do total"]))
        else:
            elementos.append(Paragraph("Sem dados no período.", styles["Normal"]))
        elementos.append(Spacer(1, 0.6 * cm))

        elementos.append(Paragraph("Mapa de Resistência", styles["Heading2"]))
        if indicadores.taxa_resistencia:
            dados_resist = [
                [
                    r.antimicrobiano,
                    str(r.total_testado),
                    str(r.total_resistente),
                    f"{r.percentual_resistente}%",
                    str(r.total_sensivel),
                    f"{r.percentual_sensivel}%",
                ]
                for r in indicadores.taxa_resistencia
            ]
            elementos.append(
                _tabela(
                    dados_resist,
                    [
                        "Antimicrobiano",
                        "Testados",
                        "Resistentes",
                        "% Resistência",
                        "Sensíveis",
                        "% Sensibilidade",
                    ],
                )
            )
        else:
            elementos.append(Paragraph("Nenhum antibiograma liberado no período.", styles["Normal"]))

        doc.build(elementos)
        return buffer.getvalue()


def _tabela(dados: list[list[str]], cabecalho: list[str]) -> Table:
    linhas = [cabecalho] + dados
    tabela = Table(linhas, hAlign="LEFT")
    tabela.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F4C81")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    return tabela
